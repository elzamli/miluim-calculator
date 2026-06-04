import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import datetime

year = 2026

def parse_ddmm(ddmm, hhmm):
    d, m = ddmm.split('/')
    h, mi = hhmm.split(':')
    return datetime(year, int(m), int(d), int(h), int(mi))

ok_pat   = re.compile(r'\[(\d+)/(\d+)\]\s+(\d+/\d+)\s+(\d+:\d+)\s+->\s+(\d+/\d+)\s+(\d+:\d+)\s+fetching\.\.\.\s+OK \((\d+) lines\)')
fail_pat = re.compile(r'\[(\d+)/(\d+)\]\s+(\d+/\d+)\s+(\d+:\d+)\s+->\s+(\d+/\d+)\s+(\d+:\d+)\s+fetching\.\.\.\s+.*FAILED')
retry_pat = re.compile(r'\[(\d+)/(\d+)\]\s+([\d-]+)\s+([\d-]+_to_[\d-]+)\s+retrying\.\.\.\s+OK \((\d+) lines\)')

rows = []
with open('run_output.txt', encoding='utf-16') as f:
    for line in f:
        m = ok_pat.search(line)
        if m:
            ds = parse_ddmm(m.group(3), m.group(4))
            de = parse_ddmm(m.group(5), m.group(6))
            rows.append({'date': ds.strftime('%Y-%m-%d'), 'from': ds.strftime('%H:%M'), 'to': de.strftime('%H:%M'), 'count': int(m.group(7)), 'status': 'OK'})
            continue
        m = fail_pat.search(line)
        if m:
            ds = parse_ddmm(m.group(3), m.group(4))
            de = parse_ddmm(m.group(5), m.group(6))
            rows.append({'date': ds.strftime('%Y-%m-%d'), 'from': ds.strftime('%H:%M'), 'to': de.strftime('%H:%M'), 'count': None, 'status': 'FAILED'})

retry_data = {}
try:
    with open('retry_output.txt', encoding='utf-16') as f:
        for line in f:
            m = retry_pat.search(line)
            if m:
                date_str = m.group(3)
                label    = m.group(4)
                cnt      = int(m.group(5))
                parts = label.split('_to_')
                hh, mm = parts[0].split('-')
                retry_data[(date_str, hh + ':' + mm)] = cnt
except (FileNotFoundError, UnicodeError):
    try:
        with open('retry_output.txt', encoding='utf-8') as f:
            for line in f:
                m = retry_pat.search(line)
                if m:
                    date_str = m.group(3)
                    label    = m.group(4)
                    cnt      = int(m.group(5))
                    parts = label.split('_to_')
                    hh, mm = parts[0].split('-')
                    retry_data[(date_str, hh + ':' + mm)] = cnt
    except FileNotFoundError:
        pass

for r in rows:
    if r['status'] == 'FAILED':
        key = (r['date'], r['from'])
        if key in retry_data:
            r['count'] = retry_data[key]
            r['status'] = 'OK (retry)'

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Logs Summary'

header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(bold=True, color='FFFFFF')
retry_fill  = PatternFill('solid', fgColor='FFF2CC')
colors_by_date = ['C6EFCE', 'BDD7EE']
date_color_map = {}
color_idx = 0

headers = ['תאריך', 'משעה', 'עד שעה', 'כמות רשומות', 'סטטוס']
ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

for r in rows:
    d = r['date']
    if d not in date_color_map:
        date_color_map[d] = colors_by_date[color_idx % 2]
        color_idx += 1
    ws.append([d, r['from'], r['to'], r['count'] if r['count'] is not None else 0, r['status']])
    row_num = ws.max_row
    fill = retry_fill if r['status'] == 'OK (retry)' else PatternFill('solid', fgColor=date_color_map[d])
    for cell in ws[row_num]:
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 10
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 14

dates = sorted(set(r['date'] for r in rows))
if len(dates) > 1:
    ws2 = wb.create_sheet('Summary by Date')
    ws2.append(['תאריך', 'סה"כ רשומות', 'מספר חלונות'])
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    for d in dates:
        day_rows = [r for r in rows if r['date'] == d]
        ws2.append([d, sum(r['count'] or 0 for r in day_rows), len(day_rows)])
    for col in ['A', 'B', 'C']:
        ws2.column_dimensions[col].width = 16

out = 'logs_summary_2906_jun.xlsx'
wb.save(out)
print('Saved: ' + out)
print('Total rows: ' + str(len(rows)))
print('Total records: ' + str(sum(r["count"] or 0 for r in rows)))
print('Dates: ' + ', '.join(sorted(set(r["date"] for r in rows))))
