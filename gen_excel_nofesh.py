import re
import os
import glob
from collections import defaultdict
from datetime import date as Date
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

BASE_DIR = r'C:\RenderLogs\NOFESH'

nofesh_pat = re.compile(
    r'\[NOFESH\] \S+ \| '
    r'aleph_plus=\d+ aleph=\d+ bet=\d+ gimel=\d+ dalet=\d+ hey=\d+ child=\w+ \| '
    r'base=\S* total=([\d.]+) source=(\S+)'
)
file_pat = re.compile(r'logs_(\d{4}-\d{2}-\d{2})_(\d{2}-\d{2})_to_(\d{2}-\d{2})\.txt')

# (date, from, to, source) -> {count, total}
data = defaultdict(lambda: {'count': 0, 'total': 0.0})

all_files = sorted(glob.glob(os.path.join(BASE_DIR, '**', 'logs_*.txt'), recursive=True))

def parse_file(filepath, date_str, from_hhmm, to_hhmm):
    for enc in ('utf-16', 'utf-8'):
        try:
            with open(filepath, encoding=enc) as f:
                for line in f:
                    m = nofesh_pat.search(line)
                    if m:
                        total  = float(m.group(1))
                        source = m.group(2)
                        data[(date_str, from_hhmm, to_hhmm, source)]['count'] += 1
                        data[(date_str, from_hhmm, to_hhmm, source)]['total'] += total
            return
        except (UnicodeError, UnicodeDecodeError):
            continue

for filepath in all_files:
    fname = os.path.basename(filepath)
    m = file_pat.match(fname)
    if not m:
        continue
    date_str  = m.group(1)
    from_hhmm = m.group(2).replace('-', ':')
    to_hhmm   = m.group(3).replace('-', ':')
    parse_file(filepath, date_str, from_hhmm, to_hhmm)

# ─── Excel ───────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()

HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF')
CENTER      = Alignment(horizontal='center')

DIRECT_FILL = PatternFill('solid', fgColor='D9D9D9')
SOURCE_PALETTE = ['C6EFCE', 'BDD7EE', 'FFE699', 'D9D2E9', 'FCE5CD', 'F4CCCC']
source_color_map = {}

def source_fill(source):
    if source == 'direct':
        return DIRECT_FILL
    if source not in source_color_map:
        idx = len(source_color_map)
        source_color_map[source] = PatternFill('solid', fgColor=SOURCE_PALETTE[idx % len(SOURCE_PALETTE)])
    return source_color_map[source]

# ── Sheet 1: פירוט לפי חלון זמן ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'פירוט לפי חלון'

headers1 = ['תאריך', 'משעה', 'עד שעה', 'source', 'חישובים', 'סכום total']
ws1.append(headers1)
for cell in ws1[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER

rows_sorted = sorted(data.keys())  # (date, from, to, source)
for (date_str, from_h, to_h, source) in rows_sorted:
    d = data[(date_str, from_h, to_h, source)]
    ws1.append([date_str, from_h, to_h, source, d['count'], round(d['total'])])
    row_num = ws1.max_row
    fill = source_fill(source)
    for cell in ws1[row_num]:
        cell.fill = fill
        cell.alignment = CENTER

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 14

# ── Sheet 2: סיכום לפי source ────────────────────────────────────────────────
source_summary = defaultdict(lambda: {'count': 0, 'total': 0.0})
for (date_str, from_h, to_h, source), d in data.items():
    source_summary[source]['count'] += d['count']
    source_summary[source]['total'] += d['total']

ws2 = wb.create_sheet('סיכום לפי source')
headers2 = ['source', 'סה"כ חישובים', 'סכום כולל', 'ממוצע total']
ws2.append(headers2)
for cell in ws2[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER

for source in sorted(source_summary.keys(), key=lambda s: -source_summary[s]['count']):
    d = source_summary[source]
    avg = round(d['total'] / d['count']) if d['count'] else 0
    ws2.append([source, d['count'], round(d['total']), avg])
    row_num = ws2.max_row
    fill = source_fill(source)
    for cell in ws2[row_num]:
        cell.fill = fill
        cell.alignment = CENTER

for col in ['A', 'B', 'C', 'D']:
    ws2.column_dimensions[col].width = 18

# ── שמירה ────────────────────────────────────────────────────────────────────
out = 'nofesh_summary.xlsx'
wb.save(out)
print(f'Saved: {out}')
print(f'Total calculations: {sum(d["count"] for d in data.values())}')
print(f'Sources: {", ".join(sorted(source_summary.keys()))}')
