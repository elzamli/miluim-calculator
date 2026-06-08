import re
import sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

RUN_LOG   = 'run_output.txt'
RETRY_LOG = 'retry_output.txt'
OUT_FILE  = 'logs_summary.xlsx'

# ── Patterns ──────────────────────────────────────────────────────────────────
# Main fetch: [N/T]  DD/MM HH:MM -> DD/MM HH:MM  fetching...  [retry N/M...]  OK (K lines)
main_ok = re.compile(
    r'\[(\d+)/(\d+)\]\s+(\d+)/(\d+)\s+(\d+:\d+)\s+->\s+(\d+)/(\d+)\s+(\d+:\d+)'
    r'\s+fetching\.\.\.(?:\s+retry \d+/\d+\.\.\.)*\s+OK \((\d+) lines\)'
)
# Retry:     [N/T]  YYYY-MM-DD HH-MM_to_HH-MM  retrying...  [retry N/M...]  OK (K lines)
retry_ok = re.compile(
    r'\[(\d+)/(\d+)\]\s+(\d{4}-\d{2}-\d{2})\s+(\d{2}-\d{2})_to_(\d{2}-\d{2})'
    r'\s+retrying\.\.\.(?:\s+retry \d+/\d+\.\.\.)*\s+OK \((\d+) lines\)'
)

def read_file(path):
    for enc in ('utf-16', 'utf-8-sig', 'utf-8'):
        try:
            with open(path, encoding=enc) as f:
                return f.read()
        except (UnicodeError, UnicodeDecodeError):
            continue
    return ''

# ── Parse main run ────────────────────────────────────────────────────────────
rows = []   # (date_str, from_hm, to_hm, lines, status)
retry_keys = set()

text = read_file(RUN_LOG)
for m in main_ok.finditer(text):
    day, mon, from_hm, end_day, end_mon, to_hm, n_lines = (
        m.group(3), m.group(4), m.group(5),
        m.group(6), m.group(7), m.group(8),
        int(m.group(9))
    )
    date_str = f'2026-{mon.zfill(2)}-{day.zfill(2)}'
    rows.append((date_str, from_hm, to_hm, n_lines, 'OK'))

# ── Parse retry run ───────────────────────────────────────────────────────────
retry_rows = []
text2 = read_file(RETRY_LOG)
for m in retry_ok.finditer(text2):
    date_str = m.group(3)
    from_hm  = m.group(4).replace('-', ':')
    to_hm    = m.group(5).replace('-', ':')
    n_lines  = int(m.group(6))
    key = (date_str, from_hm)
    retry_keys.add(key)
    retry_rows.append((date_str, from_hm, to_hm, n_lines, 'OK (retry)'))

# Merge: replace matching OK rows with OK (retry) data, append any new ones
ok_index = {(r[0], r[1]): i for i, r in enumerate(rows)}
for r in retry_rows:
    k = (r[0], r[1])
    if k in ok_index:
        rows[ok_index[k]] = r
    else:
        rows.append(r)

rows.sort(key=lambda r: (r[0], r[1]))

# ── Summary by date ───────────────────────────────────────────────────────────
date_summary = defaultdict(lambda: {'windows': 0, 'lines': 0, 'retries': 0})
for date_str, from_hm, to_hm, n_lines, status in rows:
    date_summary[date_str]['windows'] += 1
    date_summary[date_str]['lines']   += n_lines
    if status == 'OK (retry)':
        date_summary[date_str]['retries'] += 1

# ── Styles ────────────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Calibri')
RETRY_FILL   = PatternFill('solid', fgColor='FFF2CC')
DATE_FILLS   = [
    PatternFill('solid', fgColor='C6EFCE'),  # green
    PatternFill('solid', fgColor='BDD7EE'),  # blue
    PatternFill('solid', fgColor='C6EFCE'),
    PatternFill('solid', fgColor='BDD7EE'),
    PatternFill('solid', fgColor='C6EFCE'),
    PatternFill('solid', fgColor='BDD7EE'),
]
FOOTER_FILL  = PatternFill('solid', fgColor='1F4E79')
FOOTER_FONT  = Font(bold=True, color='FFFFFF', name='Calibri')
CENTER       = Alignment(horizontal='center')
RIGHT_RTL    = Alignment(horizontal='right', readingOrder=2)

wb = openpyxl.Workbook()

# ── Sheet 1: Logs Summary ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = 'סיכום לוגים'
ws1.sheet_view.rightToLeft = True

headers = ['תאריך', 'משעה', 'עד שעה', 'כמות רשומות', 'סטטוס']
ws1.append(headers)
for cell in ws1[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER

date_list = sorted(date_summary.keys())
date_color = {d: DATE_FILLS[i % len(DATE_FILLS)] for i, d in enumerate(date_list)}

for date_str, from_hm, to_hm, n_lines, status in rows:
    ws1.append([date_str, from_hm, to_hm, n_lines, status])
    row_num = ws1.max_row
    fill = RETRY_FILL if status == 'OK (retry)' else date_color.get(date_str, DATE_FILLS[0])
    for cell in ws1[row_num]:
        cell.fill = fill
        cell.alignment = CENTER

# Footer totals
total_windows = len(rows)
total_lines   = sum(r[3] for r in rows)
total_retries = sum(1 for r in rows if r[4] == 'OK (retry)')
ws1.append(['סה"כ', '', '', total_lines, f'{total_windows} חלונות ({total_retries} retry)'])
row_num = ws1.max_row
for cell in ws1[row_num]:
    cell.fill = FOOTER_FILL
    cell.font = FOOTER_FONT
    cell.alignment = CENTER

ws1.column_dimensions['A'].width = 14
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 16
ws1.column_dimensions['E'].width = 18

# ── Sheet 2: Summary by Date ──────────────────────────────────────────────────
ws2 = wb.create_sheet('סיכום לפי תאריך')
ws2.sheet_view.rightToLeft = True

headers2 = ['תאריך', 'חלונות', 'סה"כ רשומות', 'retry']
ws2.append(headers2)
for cell in ws2[1]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER

for i, date_str in enumerate(sorted(date_summary.keys())):
    d = date_summary[date_str]
    ws2.append([date_str, d['windows'], d['lines'], d['retries']])
    row_num = ws2.max_row
    fill = DATE_FILLS[i % len(DATE_FILLS)]
    for cell in ws2[row_num]:
        cell.fill = fill
        cell.alignment = CENTER

ws2.append(['סה"כ',
            sum(d['windows'] for d in date_summary.values()),
            sum(d['lines']   for d in date_summary.values()),
            sum(d['retries'] for d in date_summary.values())])
row_num = ws2.max_row
for cell in ws2[row_num]:
    cell.fill = FOOTER_FILL
    cell.font = FOOTER_FONT
    cell.alignment = CENTER

for col in ['A', 'B', 'C', 'D']:
    ws2.column_dimensions[col].width = 16

wb.save(OUT_FILE)
print(f'Saved: {OUT_FILE}')
print(f'Total windows : {total_windows}')
print(f'Total lines   : {total_lines}')
print(f'Retry windows : {total_retries}')
print(f'Dates covered : {", ".join(sorted(date_summary.keys()))}')
